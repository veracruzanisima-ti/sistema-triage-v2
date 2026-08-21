"""Exportador DIF v1: renderiza el modelo interno sin incorporar reglas propias."""

import re
from dataclasses import dataclass
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from triage.comercial.modelos import EstadoComercial
from triage.cotizaciones.modelos import Cotizacion
from triage.cotizaciones.servicio import obtener_cotizacion
from triage.normalizacion.servicio import resumen_normalizacion_cotizacion
from triage.revision_final.servicio import ProductoPreCierre, listar_precierre

_MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_FORMATO_MONEDA = '$#,##0.00'
_GUION = "—"
_PREFIJOS_FORMULA = ("=", "+", "-", "@")


class ErrorExportacionDif(ValueError):
    """El modelo interno todavía no cumple las condiciones para emitir DIF."""


@dataclass(frozen=True)
class ExportacionDif:
    contenido: bytes
    nombre_archivo: str
    mime_type: str = _MIME_XLSX


def _texto(valor: object | None) -> str:
    return " ".join(str(valor or "").split())


def _texto_excel(valor: object | None) -> str:
    """Evita que texto controlado por usuarios se interprete como fórmula al abrir el XLSX."""

    limpio = _texto(valor)
    if limpio.startswith(_PREFIJOS_FORMULA):
        return "'" + limpio
    return limpio


def _nombre_archivo(cotizacion: Cotizacion) -> str:
    base = _texto(cotizacion.referencia) or cotizacion.id[:8]
    seguro = re.sub(r"[^A-Za-z0-9._-]+", "_", base).strip("._-") or cotizacion.id[:8]
    return f"Cotizacion_DIF_{seguro[:80]}.xlsx"


def _validar_exportable(
    sesion: Session,
    *,
    cotizacion: Cotizacion,
    productos: list[ProductoPreCierre],
) -> None:
    resumen = resumen_normalizacion_cotizacion(sesion, cotizacion.id)
    if resumen.total == 0:
        raise ErrorExportacionDif("La cotización no tiene partidas revisadas para exportar.")
    if resumen.preparados != resumen.total:
        faltantes = resumen.total - resumen.preparados
        raise ErrorExportacionDif(
            f"Faltan {faltantes} partida(s) por preparar antes de exportar DIF."
        )
    if len(productos) != resumen.total:
        raise ErrorExportacionDif(
            "La cotización no pudo consolidar todas sus partidas preparadas."
        )

    pendientes: list[str] = []
    for item in productos:
        if item.decision_comercial.estado == EstadoComercial.NO_SE_COTIZA:
            continue
        if item.referencia is None:
            pendientes.append("referencia estable")
        if item.precio_venta is None:
            pendientes.append("precio unitario final sin IVA")
        if item.validacion_fiscal is None:
            pendientes.append("validación fiscal")
        if item.calculo_venta is None:
            pendientes.append("cálculo final")
    if pendientes:
        tipos = ", ".join(dict.fromkeys(pendientes))
        raise ErrorExportacionDif(
            "La cotización no es emitible todavía. Pendiente: " + tipos + "."
        )


def _fila_dif(numero: int, item: ProductoPreCierre) -> list[object]:
    producto = item.producto
    partida = producto.partida
    documento = producto.documento
    normalizacion = producto.normalizacion
    folios = ", ".join(str(folio) for folio in documento.folios)

    if item.decision_comercial.estado == EstadoComercial.NO_SE_COTIZA:
        return [
            numero,
            _texto_excel(documento.memorandum),
            _texto_excel(folios),
            _texto_excel(documento.municipio),
            _texto_excel(normalizacion.producto or partida.producto_solicitado),
            _texto_excel(normalizacion.presentacion or partida.presentacion_solicitada),
            float(partida.cantidad) if partida.cantidad is not None else "",
            _texto_excel(partida.unidad_medida),
            _texto_excel(normalizacion.marca or partida.marca_solicitada),
            "NO SE COTIZA",
            _texto_excel(item.decision_comercial.motivo),
            _GUION,
            _GUION,
            _GUION,
            _GUION,
            _GUION,
            _GUION,
            _GUION,
            _GUION,
        ]

    calculo = item.calculo_venta
    validacion = item.validacion_fiscal
    referencia = item.referencia
    precio_venta = item.precio_venta
    if calculo is None or validacion is None or referencia is None or precio_venta is None:
        raise ErrorExportacionDif(
            "Una partida cotizable perdió su consolidación antes de exportar."
        )

    return [
        numero,
        _texto_excel(documento.memorandum),
        _texto_excel(folios),
        _texto_excel(documento.municipio),
        _texto_excel(normalizacion.producto or partida.producto_solicitado),
        _texto_excel(normalizacion.presentacion or partida.presentacion_solicitada),
        float(partida.cantidad) if partida.cantidad is not None else "",
        _texto_excel(partida.unidad_medida),
        _texto_excel(normalizacion.marca or partida.marca_solicitada),
        "COTIZABLE",
        "",
        float(calculo.precio_unitario_sin_iva),
        float(calculo.subtotal),
        float(calculo.iva),
        float(calculo.total),
        _texto_excel(validacion.etiqueta),
        _texto_excel(referencia.proveedor),
        _texto_excel(referencia.fuente),
        _texto_excel(calculo.origen_precio),
    ]


def _crear_libro(cotizacion: Cotizacion, productos: list[ProductoPreCierre]) -> bytes:
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Cotización DIF"
    hoja.append(["COTIZACIÓN DIF · VERACRUZANÍSIMA"])
    hoja.append(["Referencia", _texto_excel(cotizacion.referencia or "Sin referencia")])
    hoja.append(
        [
            "Trazabilidad",
            "Los importes parten del precio unitario final sin IVA confirmado por una persona "
            "y del tratamiento fiscal validado. El exportador no calcula margen ni tasa.",
        ]
    )
    hoja.append([])

    encabezados = [
        "N",
        "Memorándum",
        "Folio(s)",
        "Municipio",
        "Producto",
        "Presentación",
        "Cantidad",
        "Unidad de medida",
        "Marca",
        "Resultado comercial",
        "Motivo",
        "Precio unitario s/IVA",
        "Subtotal",
        "IVA",
        "Total",
        "Tratamiento fiscal",
        "Proveedor de referencia",
        "Fuente",
        "Origen del precio",
    ]
    hoja.append(encabezados)

    for numero, item in enumerate(productos, start=1):
        hoja.append(_fila_dif(numero, item))

    hoja.freeze_panes = "A6"
    hoja.auto_filter.ref = f"A5:S{hoja.max_row}"
    hoja["A1"].font = Font(bold=True, size=14)
    hoja["A3"].font = Font(bold=True)
    for celda in hoja[5]:
        celda.font = Font(bold=True)
        celda.alignment = Alignment(wrap_text=True, vertical="top")
    for fila in hoja.iter_rows(min_row=6):
        for celda in fila:
            celda.alignment = Alignment(vertical="top", wrap_text=True)
    for fila in range(6, hoja.max_row + 1):
        for columna in range(12, 16):
            celda = hoja.cell(row=fila, column=columna)
            if isinstance(celda.value, (int, float)):
                celda.number_format = _FORMATO_MONEDA

    anchos = {
        1: 6,
        2: 24,
        3: 22,
        4: 20,
        5: 32,
        6: 30,
        7: 12,
        8: 18,
        9: 22,
        10: 20,
        11: 34,
        12: 20,
        13: 16,
        14: 14,
        15: 16,
        16: 20,
        17: 24,
        18: 42,
        19: 44,
    }
    for indice, ancho in anchos.items():
        hoja.column_dimensions[get_column_letter(indice)].width = ancho

    salida = BytesIO()
    libro.save(salida)
    libro.close()
    return salida.getvalue()


def generar_exportacion_dif(sesion: Session, cotizacion_id: str) -> ExportacionDif:
    """Genera DIF sólo cuando precio final y tratamiento fiscal ya fueron confirmados."""

    cotizacion = obtener_cotizacion(sesion, cotizacion_id)
    if cotizacion is None:
        raise LookupError("Cotización no encontrada")
    productos = listar_precierre(sesion, cotizacion_id)
    _validar_exportable(sesion, cotizacion=cotizacion, productos=productos)
    return ExportacionDif(
        contenido=_crear_libro(cotizacion, productos),
        nombre_archivo=_nombre_archivo(cotizacion),
    )
