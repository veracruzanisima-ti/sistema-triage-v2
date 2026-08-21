"""Importación transaccional y resolución conservadora de identidad COFEPRIS."""

from __future__ import annotations

import re
from dataclasses import dataclass
from hashlib import sha256
from io import BytesIO
from pathlib import Path
from uuid import uuid4
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import delete, func, insert, select
from sqlalchemy.orm import Session

from triage.proveedores.cofepris_modelos import ImportacionCofepris, RegistroCofepris
from triage.proveedores.coincidencia_catalogo import normalizar_texto

HOJA_REGISTROS_COFEPRIS = "Registros_Medicamentos"
ESTADO_VIGENTE = "VIGENTE"
_MAX_XLSX_BYTES = 50 * 1024 * 1024
_MAX_DENOMINACION_NORMALIZADA = 500
_MAX_NUMERO_REGISTRO = 255

_COLUMNAS = {
    "numero_registro": "Número de Registro",
    "denominacion_distintiva": "Denominación Distintiva",
    "estado": "Estado",
    "forma_farmaceutica": "Forma Farmaceutica",
    "denominacion_generica": "Denominacion Generica",
    "via_administracion": "Vista Administración",
    "tipo_medicamento": "Tipo Medicamento",
    "presentacion": "Presentación",
    "cantidad": "Cantidad",
    "fraccion_sanitaria": "Fracción",
    "sustancia_quimica": "Sustancia Química",
    "titular": "Titular",
    "fecha_emision": "Fecha Emisión",
}
_OBLIGATORIAS_POR_REGISTRO = (
    "numero_registro",
    "denominacion_distintiva",
    "denominacion_generica",
    "estado",
)


class ErrorImportacionCofepris(ValueError):
    """El XLSX es inválido y no debe sustituir el snapshot vigente."""


@dataclass(frozen=True)
class EvidenciaIdentidadCofepris:
    """Copia autocontenida de la identidad pública usada al aceptar un candidato."""

    fuente: str
    numero_registro: str
    denominacion_distintiva: str
    denominacion_generica: str
    estado: str
    importacion_id: str
    sha256_importacion: str
    numeros_registro: tuple[str, ...] = ()

    def como_json(self) -> dict[str, object]:
        evidencia: dict[str, object] = {
            "fuente": self.fuente,
            "numero_registro": self.numero_registro,
            "denominacion_distintiva": self.denominacion_distintiva,
            "denominacion_generica": self.denominacion_generica,
            "estado": self.estado,
            "importacion_id": self.importacion_id,
            "sha256_importacion": self.sha256_importacion,
        }
        if len(self.numeros_registro) > 1:
            evidencia["numeros_registro"] = list(self.numeros_registro)
            evidencia["registros_coincidentes"] = len(self.numeros_registro)
        return evidencia


@dataclass(frozen=True)
class ResultadoParseoCofepris:
    """Filas oficiales y señales de calidad que no deben ocultarse."""

    registros: list[dict[str, object]]
    registros_sin_identidad_util: int
    numeros_registro_duplicados: int


def _limpiar(valor: object) -> str | None:
    if valor is None:
        return None
    limpio = " ".join(str(valor).split())
    return limpio or None


def _nombre_archivo(nombre: str) -> str:
    limpio = Path(nombre or "").name.strip()
    if not limpio or not limpio.lower().endswith(".xlsx"):
        raise ErrorImportacionCofepris("El catálogo COFEPRIS debe ser un archivo .xlsx")
    if len(limpio) > 255:
        raise ErrorImportacionCofepris("El nombre del archivo excede 255 caracteres")
    return limpio


def _componentes_genericos(valor: str) -> list[str]:
    normalizado = normalizar_texto(valor)
    componentes = [
        componente.strip()
        for componente in re.split(r"\s*(?:\+|/|;|,|\bY\b)\s*", normalizado)
        if componente.strip()
    ]
    return list(dict.fromkeys(componentes))


def _parsear_xlsx(datos: bytes) -> ResultadoParseoCofepris:
    if not datos:
        raise ErrorImportacionCofepris("El catálogo COFEPRIS está vacío")
    if len(datos) > _MAX_XLSX_BYTES:
        raise ErrorImportacionCofepris("El catálogo COFEPRIS excede el tamaño permitido")

    try:
        libro = load_workbook(BytesIO(datos), read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as error:
        raise ErrorImportacionCofepris("El archivo no es un XLSX válido") from error

    try:
        if HOJA_REGISTROS_COFEPRIS not in libro.sheetnames:
            raise ErrorImportacionCofepris(
                f"El archivo no contiene la hoja {HOJA_REGISTROS_COFEPRIS}"
            )
        hoja = libro[HOJA_REGISTROS_COFEPRIS]
        filas = hoja.iter_rows(values_only=True)
        encabezados_crudos = next(filas, None)
        if encabezados_crudos is None:
            raise ErrorImportacionCofepris("La hoja de registros está vacía")
        encabezados = {
            _limpiar(valor): indice
            for indice, valor in enumerate(encabezados_crudos)
            if _limpiar(valor)
        }
        faltantes = [columna for columna in _COLUMNAS.values() if columna not in encabezados]
        if faltantes:
            raise ErrorImportacionCofepris(
                "Faltan columnas obligatorias: " + ", ".join(faltantes)
            )

        registros: list[dict[str, object]] = []
        numeros_vistos: set[str] = set()
        numeros_duplicados: set[str] = set()
        registros_sin_identidad_util = 0
        for numero_fila, fila in enumerate(filas, start=2):
            if not any(_limpiar(valor) for valor in fila):
                continue
            valores = {
                campo: _limpiar(fila[encabezados[columna]])
                if encabezados[columna] < len(fila)
                else None
                for campo, columna in _COLUMNAS.items()
            }
            for campo in _OBLIGATORIAS_POR_REGISTRO:
                if valores[campo] is None:
                    raise ErrorImportacionCofepris(
                        f"La fila {numero_fila} no contiene {_COLUMNAS[campo]}"
                    )

            numero_registro = str(valores["numero_registro"])
            clave_registro = numero_registro.casefold()
            if clave_registro in numeros_vistos:
                numeros_duplicados.add(clave_registro)
            numeros_vistos.add(clave_registro)
            if len(numero_registro) > _MAX_NUMERO_REGISTRO:
                raise ErrorImportacionCofepris(
                    f"El Número de Registro de la fila {numero_fila} excede 255 caracteres"
                )

            distintiva_normalizada = normalizar_texto(
                str(valores["denominacion_distintiva"])
            )
            if len(distintiva_normalizada) > _MAX_DENOMINACION_NORMALIZADA:
                raise ErrorImportacionCofepris(
                    f"La denominación distintiva de la fila {numero_fila} excede el límite"
                )
            componentes = _componentes_genericos(str(valores["denominacion_generica"]))
            if not distintiva_normalizada or not componentes:
                registros_sin_identidad_util += 1

            registros.append(
                {
                    **valores,
                    "id": str(uuid4()),
                    "numero_registro": numero_registro,
                    "denominacion_distintiva_normalizada": distintiva_normalizada,
                    "componentes_genericos_normalizados": componentes,
                    "estado": normalizar_texto(str(valores["estado"])),
                }
            )
    finally:
        libro.close()

    if not registros:
        raise ErrorImportacionCofepris("El catálogo COFEPRIS no contiene registros")
    return ResultadoParseoCofepris(
        registros=registros,
        registros_sin_identidad_util=registros_sin_identidad_util,
        numeros_registro_duplicados=len(numeros_duplicados),
    )


def importar_snapshot_cofepris(
    sesion: Session,
    *,
    usuario_id: str,
    nombre_archivo: str,
    datos: bytes,
) -> ImportacionCofepris:
    """Valida todo el XLSX antes de reemplazar el snapshot en una transacción."""

    archivo = _nombre_archivo(nombre_archivo)
    resultado = _parsear_xlsx(datos)
    registros = resultado.registros
    importacion_id = str(uuid4())
    importacion = ImportacionCofepris(
        id=importacion_id,
        cargada_por_usuario_id=usuario_id,
        archivo=archivo,
        sha256=sha256(datos).hexdigest(),
        registros_cargados=len(registros),
        registros_vigentes=sum(
            registro["estado"] == ESTADO_VIGENTE for registro in registros
        ),
        registros_sin_identidad_util=resultado.registros_sin_identidad_util,
        numeros_registro_duplicados=resultado.numeros_registro_duplicados,
    )
    filas = [{**registro, "importacion_id": importacion_id} for registro in registros]

    try:
        sesion.add(importacion)
        sesion.flush()
        sesion.execute(delete(RegistroCofepris))
        sesion.execute(insert(RegistroCofepris), filas)
        sesion.commit()
    except Exception:
        sesion.rollback()
        raise
    sesion.refresh(importacion)
    return importacion


def ultima_importacion_cofepris(sesion: Session) -> ImportacionCofepris | None:
    return sesion.scalar(
        select(ImportacionCofepris)
        .order_by(ImportacionCofepris.cargada_en.desc())
        .limit(1)
    )


def total_registros_cofepris(sesion: Session) -> int:
    return int(sesion.scalar(select(func.count()).select_from(RegistroCofepris)) or 0)


def _perfil_registro(registro: RegistroCofepris) -> tuple[str, str, str]:
    """Distingue registros/presentaciones reales de duplicados equivalentes del snapshot."""

    return (
        normalizar_texto(registro.forma_farmaceutica),
        normalizar_texto(registro.presentacion),
        normalizar_texto(registro.cantidad),
    )


def resolver_identidad_cofepris(
    sesion: Session,
    *,
    producto_solicitado: str,
    producto_observado: str,
) -> EvidenciaIdentidadCofepris | None:
    """Resuelve marca inicial exacta si uno o varios registros VIGENTES convergen."""

    observado = normalizar_texto(producto_observado)
    solicitado = normalizar_texto(producto_solicitado)
    tokens = observado.split()[:20]
    if not tokens or not solicitado:
        return None
    prefijos = [" ".join(tokens[:fin]) for fin in range(1, len(tokens) + 1)]
    registros = list(
        sesion.scalars(
            select(RegistroCofepris).where(
                RegistroCofepris.estado == ESTADO_VIGENTE,
                RegistroCofepris.denominacion_distintiva_normalizada.in_(prefijos),
            )
        )
    )
    if not registros:
        return None
    longitud_maxima = max(
        len(registro.denominacion_distintiva_normalizada.split()) for registro in registros
    )
    coincidencias = [
        registro
        for registro in registros
        if len(registro.denominacion_distintiva_normalizada.split()) == longitud_maxima
    ]

    if any(
        len(registro.componentes_genericos_normalizados) != 1
        or registro.componentes_genericos_normalizados[0] != solicitado
        for registro in coincidencias
    ):
        return None

    numeros = [registro.numero_registro for registro in coincidencias]
    if len({numero.casefold() for numero in numeros}) != len(numeros):
        return None
    perfiles = {_perfil_registro(registro) for registro in coincidencias}
    if len(coincidencias) > 1 and len(perfiles) == 1:
        return None

    registro = sorted(coincidencias, key=lambda item: item.numero_registro.casefold())[0]
    importacion = sesion.get(ImportacionCofepris, registro.importacion_id)
    if importacion is None or any(
        coincidencia.importacion_id != importacion.id for coincidencia in coincidencias
    ):
        return None
    numeros_ordenados = tuple(sorted(numeros, key=str.casefold))
    return EvidenciaIdentidadCofepris(
        fuente="COFEPRIS",
        numero_registro=registro.numero_registro,
        denominacion_distintiva=registro.denominacion_distintiva,
        denominacion_generica=registro.denominacion_generica,
        estado=registro.estado,
        importacion_id=importacion.id,
        sha256_importacion=importacion.sha256,
        numeros_registro=numeros_ordenados,
    )