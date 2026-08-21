"""Prueba de humo del flujo crítico que se validará manualmente en el piloto DIF."""

from decimal import Decimal
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from triage.comercial.modelos import EstadoComercial
from triage.comercial.precios_venta_servicio import registrar_precio_venta
from triage.comercial.servicio import registrar_decision_comercial
from triage.cotizaciones.modelos import Cotizacion
from triage.documentos.modelos import Documento, EstadoDocumento, PartidaDocumento
from triage.fiscal.modelos import TratamientoIVA
from triage.fiscal.servicio import registrar_validacion_fiscal
from triage.historico.decisiones_modelos import RolDecisionPrecio
from triage.historico.decisiones_servicio import registrar_decision_precio
from triage.historico.servicio import crear_observacion_precio
from triage.normalizacion.modelos import NormalizacionPartida
from triage.proveedores.base import SolicitudProveedor
from triage.proveedores.coincidencia_catalogo import CandidatoCatalogo, evaluar_candidato
from triage.usuarios.modelos import Usuario


def test_piloto_dif_integra_matcher_precio_final_no_cotiza_y_excel(cliente: TestClient):
    """Protege el recorrido de negocio principal sin llamar proveedores o IA externos."""

    with cliente.app.state.fabrica_sesiones() as sesion:
        usuario = sesion.scalar(select(Usuario).limit(1))
        assert usuario is not None

        cotizacion = Cotizacion(
            referencia="PILOTO-DIF-E2E",
            referencia_fijada_manual=True,
        )
        sesion.add(cotizacion)
        sesion.flush()
        documento = Documento(
            cotizacion_id=cotizacion.id,
            nombre_original="piloto-dif.pdf",
            mime_type="application/pdf",
            tamano_bytes=100,
            sha256="9" * 64,
            estado=EstadoDocumento.REVISADO.value,
            memorandum="DAIS/PILOTO/001/2026",
            folios=["PILOTO-001"],
            municipio="Xalapa",
        )
        sesion.add(documento)
        sesion.flush()

        dapagliflozina = PartidaDocumento(
            documento_id=documento.id,
            orden=1,
            producto_solicitado="DAPAGLIFLOZINA",
            concentracion="10 mg",
            forma_farmaceutica_dispositivo="tabletas",
            presentacion_solicitada="Caja con 28 tabletas de 10 mg",
            cantidad=Decimal("2"),
            unidad_medida="cajas",
        )
        no_cotizable = PartidaDocumento(
            documento_id=documento.id,
            orden=2,
            producto_solicitado="PRODUCTO NO COTIZABLE DE PRUEBA",
            presentacion_solicitada="Pieza",
            cantidad=Decimal("1"),
            unidad_medida="pieza",
        )
        sesion.add_all([dapagliflozina, no_cotizable])
        sesion.flush()
        sesion.add_all(
            [
                NormalizacionPartida(
                    partida_documento_id=dapagliflozina.id,
                    producto="DAPAGLIFLOZINA",
                    concentracion="10 mg",
                    forma_dispositivo="tabletas",
                    presentacion="Caja con 28 tabletas de 10 mg",
                    confirmada_por_usuario_id=usuario.id,
                ),
                NormalizacionPartida(
                    partida_documento_id=no_cotizable.id,
                    producto="PRODUCTO NO COTIZABLE DE PRUEBA",
                    presentacion="Pieza",
                    confirmada_por_usuario_id=usuario.id,
                ),
            ]
        )
        sesion.commit()

        evaluacion = evaluar_candidato(
            SolicitudProveedor(
                partida_documento_id=dapagliflozina.id,
                producto="DAPAGLIFLOZINA",
                marca=None,
                concentracion="10 mg",
                forma_dispositivo="tabletas",
                presentacion="Caja con 28 tabletas de 10 mg",
            ),
            CandidatoCatalogo(
                descripcion="Dapagliflozina 10 mg 28 Tabs Marca del Ahorro",
                precio_observado=Decimal("762"),
                stock=1,
                fuente="https://proveedor.example/dapagliflozina",
            ),
        )
        assert evaluacion.coincide is True
        assert evaluacion.motivos == ()

        referencia = crear_observacion_precio(
            sesion,
            cotizacion_id=cotizacion.id,
            partida_documento_id=dapagliflozina.id,
            usuario_id=usuario.id,
            proveedor="Proveedor piloto",
            fuente="https://proveedor.example/dapagliflozina",
            precio_antes_iva=Decimal("762"),
            iva_porcentaje=Decimal("0"),
            precio_total=Decimal("762"),
            es_promocion=False,
            condiciones_promocion=None,
            disponibilidad="Disponible",
            entrega_viable=True,
        )
        registrar_decision_precio(
            sesion,
            cotizacion_id=cotizacion.id,
            partida_id=dapagliflozina.id,
            usuario_id=usuario.id,
            rol=RolDecisionPrecio.REFERENCIA_ESTABLE,
            observacion_id=referencia.id,
        )
        registrar_validacion_fiscal(
            sesion,
            cotizacion_id=cotizacion.id,
            partida_id=dapagliflozina.id,
            usuario_id=usuario.id,
            tratamiento_iva=TratamientoIVA.TASA,
            iva_porcentaje=Decimal("0"),
            observacion=None,
        )
        registrar_precio_venta(
            sesion,
            cotizacion_id=cotizacion.id,
            partida_id=dapagliflozina.id,
            usuario_id=usuario.id,
            precio_unitario_sin_iva=Decimal("850"),
            fuente_comercial="Autorización comercial del piloto",
            observacion="Precio final distinto de la referencia de adquisición",
        )
        registrar_decision_comercial(
            sesion,
            cotizacion_id=cotizacion.id,
            partida_id=no_cotizable.id,
            usuario_id=usuario.id,
            estado=EstadoComercial.NO_SE_COTIZA,
            motivo="Decisión comercial validada durante el piloto",
        )
        cotizacion_id = cotizacion.id

    revision = cliente.get(f"/cotizaciones/{cotizacion_id}/revision-final")
    assert revision.status_code == 200
    assert "Lista para generar el DIF" in revision.text
    assert "DAPAGLIFLOZINA" in revision.text
    assert "NO SE COTIZA" in revision.text
    assert "850.00" in revision.text

    descarga = cliente.get(f"/cotizaciones/{cotizacion_id}/exportaciones/dif.xlsx")
    assert descarga.status_code == 200
    assert descarga.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    libro = load_workbook(BytesIO(descarga.content), data_only=True)
    hoja = libro["Cotización DIF"]
    try:
        assert hoja["E6"].value == "DAPAGLIFLOZINA"
        assert hoja["J6"].value == "COTIZABLE"
        assert hoja["L6"].value == 850
        assert hoja["M6"].value == 1700
        assert hoja["N6"].value == 0
        assert hoja["O6"].value == 1700
        assert hoja["Q6"].value == "Proveedor piloto"
        assert hoja["S6"].value == "Autorización comercial del piloto"

        assert hoja["E7"].value == "PRODUCTO NO COTIZABLE DE PRUEBA"
        assert hoja["J7"].value == "NO SE COTIZA"
        assert hoja["K7"].value == "Decisión comercial validada durante el piloto"
        assert [hoja.cell(row=7, column=columna).value for columna in range(12, 16)] == [
            "—",
            "—",
            "—",
            "—",
        ]
    finally:
        libro.close()
