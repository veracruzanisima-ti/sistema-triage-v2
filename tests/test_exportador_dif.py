"""Regresiones del Exportador DIF v1 y su condición de emitibilidad."""

from decimal import Decimal
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from triage.comercial.modelos import EstadoComercial
from triage.comercial.precio_venta_servicio import registrar_precio_venta
from triage.comercial.servicio import registrar_decision_comercial
from triage.cotizaciones.modelos import Cotizacion
from triage.documentos.modelos import Documento, EstadoDocumento, PartidaDocumento
from triage.exportadores.dif import ErrorExportacionDif, generar_exportacion_dif
from triage.fiscal.modelos import TratamientoIVA
from triage.fiscal.servicio import registrar_validacion_fiscal
from triage.historico.decisiones_modelos import RolDecisionPrecio
from triage.historico.decisiones_servicio import registrar_decision_precio
from triage.historico.servicio import crear_observacion_precio
from triage.normalizacion.modelos import NormalizacionPartida
from triage.usuarios.modelos import Usuario


def _crear_producto(sesion, *, referencia: str = "DIF-001"):
    usuario = sesion.scalar(select(Usuario).limit(1))
    assert usuario is not None

    cotizacion = Cotizacion(referencia=referencia, referencia_fijada_manual=True)
    sesion.add(cotizacion)
    sesion.flush()
    documento = Documento(
        cotizacion_id=cotizacion.id,
        nombre_original="dif.pdf",
        mime_type="application/pdf",
        tamano_bytes=10,
        sha256="d" * 64,
        estado=EstadoDocumento.REVISADO.value,
        memorandum="DAIS/SSMA/123/2026",
        folios=["REC-001"],
        municipio="Xalapa",
    )
    sesion.add(documento)
    sesion.flush()
    partida = PartidaDocumento(
        documento_id=documento.id,
        orden=1,
        producto_solicitado="PARACETAMOL",
        marca_solicitada="MARCA PRUEBA",
        concentracion="500 mg",
        forma_farmaceutica_dispositivo="tabletas",
        presentacion_solicitada="Caja con 10 tabletas",
        cantidad=Decimal("2"),
        unidad_medida="cajas",
    )
    sesion.add(partida)
    sesion.flush()
    sesion.add(
        NormalizacionPartida(
            partida_documento_id=partida.id,
            producto="PARACETAMOL",
            marca="MARCA PRUEBA",
            concentracion="500 mg",
            forma_dispositivo="tabletas",
            presentacion="Caja con 10 tabletas",
            confirmada_por_usuario_id=usuario.id,
        )
    )
    sesion.commit()
    return cotizacion, partida, usuario


def _seleccionar_referencia(sesion, *, cotizacion, partida, usuario):
    referencia = crear_observacion_precio(
        sesion,
        cotizacion_id=cotizacion.id,
        partida_documento_id=partida.id,
        usuario_id=usuario.id,
        proveedor="Proveedor estable",
        fuente="https://proveedor.example/producto",
        precio_antes_iva=Decimal("100"),
        iva_porcentaje=Decimal("16"),
        precio_total=Decimal("116"),
        es_promocion=False,
        condiciones_promocion=None,
        disponibilidad="Disponible",
        entrega_viable=True,
    )
    registrar_decision_precio(
        sesion,
        cotizacion_id=cotizacion.id,
        partida_id=partida.id,
        usuario_id=usuario.id,
        rol=RolDecisionPrecio.REFERENCIA_ESTABLE,
        observacion_id=referencia.id,
    )
    return referencia


def _validar_fiscal(sesion, *, cotizacion, partida, usuario):
    registrar_validacion_fiscal(
        sesion,
        cotizacion_id=cotizacion.id,
        partida_id=partida.id,
        usuario_id=usuario.id,
        tratamiento_iva=TratamientoIVA.TASA,
        iva_porcentaje=Decimal("16"),
        observacion=None,
    )


def _hacer_emitible(sesion, *, cotizacion, partida, usuario):
    _seleccionar_referencia(
        sesion,
        cotizacion=cotizacion,
        partida=partida,
        usuario=usuario,
    )
    registrar_precio_venta(
        sesion,
        cotizacion_id=cotizacion.id,
        partida_id=partida.id,
        usuario_id=usuario.id,
        precio_unitario_sin_iva=Decimal("150"),
        observacion="Precio autorizado para la prueba",
    )
    _validar_fiscal(
        sesion,
        cotizacion=cotizacion,
        partida=partida,
        usuario=usuario,
    )


def test_dif_bloquea_si_falta_validacion_fiscal(cliente: TestClient):
    with cliente.app.state.fabrica_sesiones() as sesion:
        cotizacion, partida, usuario = _crear_producto(sesion, referencia="DIF-FISCAL-PENDIENTE")
        _seleccionar_referencia(
            sesion,
            cotizacion=cotizacion,
            partida=partida,
            usuario=usuario,
        )
        registrar_precio_venta(
            sesion,
            cotizacion_id=cotizacion.id,
            partida_id=partida.id,
            usuario_id=usuario.id,
            precio_unitario_sin_iva=Decimal("150"),
            observacion=None,
        )

        with pytest.raises(ErrorExportacionDif, match="validación fiscal"):
            generar_exportacion_dif(sesion, cotizacion.id)


def test_dif_bloquea_si_falta_precio_final(cliente: TestClient):
    with cliente.app.state.fabrica_sesiones() as sesion:
        cotizacion, partida, usuario = _crear_producto(sesion, referencia="DIF-PRECIO-PENDIENTE")
        _seleccionar_referencia(
            sesion,
            cotizacion=cotizacion,
            partida=partida,
            usuario=usuario,
        )
        _validar_fiscal(
            sesion,
            cotizacion=cotizacion,
            partida=partida,
            usuario=usuario,
        )

        with pytest.raises(ErrorExportacionDif, match="precio unitario final sin IVA"):
            generar_exportacion_dif(sesion, cotizacion.id)


def test_dif_exporta_precio_final_confirmado_y_no_costo_proveedor(cliente: TestClient):
    with cliente.app.state.fabrica_sesiones() as sesion:
        cotizacion, partida, usuario = _crear_producto(sesion, referencia="DIF-VALIDADA")
        _hacer_emitible(
            sesion,
            cotizacion=cotizacion,
            partida=partida,
            usuario=usuario,
        )

        exportacion = generar_exportacion_dif(sesion, cotizacion.id)

    libro = load_workbook(BytesIO(exportacion.contenido), data_only=True)
    hoja = libro["Cotización DIF"]
    try:
        assert exportacion.nombre_archivo == "Cotizacion_DIF_DIF-VALIDADA.xlsx"
        assert hoja["A1"].value == "COTIZACIÓN DIF · VERACRUZANÍSIMA"
        assert "precio unitario final sin IVA confirmado" in hoja["B3"].value
        assert hoja["E6"].value == "PARACETAMOL"
        assert hoja["J6"].value == "COTIZABLE"
        assert hoja["L6"].value == 150
        assert hoja["M6"].value == 300
        assert hoja["N6"].value == 48
        assert hoja["O6"].value == 348
        assert hoja["P6"].value == "IVA 16.00%"
        assert hoja["Q6"].value == "Proveedor estable"
        assert hoja["S6"].value == "Precio unitario final confirmado manualmente"
    finally:
        libro.close()


def test_dif_conserva_no_se_cotiza_y_deja_importes_en_guion(cliente: TestClient):
    with cliente.app.state.fabrica_sesiones() as sesion:
        cotizacion, partida, usuario = _crear_producto(sesion, referencia="DIF-NO-COTIZA")
        registrar_decision_comercial(
            sesion,
            cotizacion_id=cotizacion.id,
            partida_id=partida.id,
            usuario_id=usuario.id,
            estado=EstadoComercial.NO_SE_COTIZA,
            motivo="Restricción comercial validada para la prueba",
        )

        exportacion = generar_exportacion_dif(sesion, cotizacion.id)

    libro = load_workbook(BytesIO(exportacion.contenido), data_only=True)
    hoja = libro["Cotización DIF"]
    try:
        assert hoja["J6"].value == "NO SE COTIZA"
        assert hoja["K6"].value == "Restricción comercial validada para la prueba"
        assert [hoja.cell(row=6, column=columna).value for columna in range(12, 16)] == [
            "—",
            "—",
            "—",
            "—",
        ]
    finally:
        libro.close()


def test_dif_neutraliza_texto_que_excel_podria_ejecutar_como_formula(cliente: TestClient):
    with cliente.app.state.fabrica_sesiones() as sesion:
        cotizacion, partida, usuario = _crear_producto(sesion, referencia="DIF-SEGURIDAD")
        registrar_decision_comercial(
            sesion,
            cotizacion_id=cotizacion.id,
            partida_id=partida.id,
            usuario_id=usuario.id,
            estado=EstadoComercial.NO_SE_COTIZA,
            motivo="=HYPERLINK(\"https://example.invalid\",\"abrir\")",
        )
        exportacion = generar_exportacion_dif(sesion, cotizacion.id)

    libro = load_workbook(BytesIO(exportacion.contenido), data_only=False)
    hoja = libro["Cotización DIF"]
    try:
        assert hoja["K6"].data_type == "s"
        assert hoja["K6"].value.startswith("'=")
    finally:
        libro.close()


def test_ruta_dif_responde_conflicto_mientras_falte_precio_final(cliente: TestClient):
    with cliente.app.state.fabrica_sesiones() as sesion:
        cotizacion, partida, usuario = _crear_producto(sesion, referencia="DIF-RUTA-PENDIENTE")
        _seleccionar_referencia(
            sesion,
            cotizacion=cotizacion,
            partida=partida,
            usuario=usuario,
        )
        _validar_fiscal(
            sesion,
            cotizacion=cotizacion,
            partida=partida,
            usuario=usuario,
        )
        cotizacion_id = cotizacion.id

    respuesta = cliente.get(f"/cotizaciones/{cotizacion_id}/exportaciones/dif.xlsx")

    assert respuesta.status_code == 409
    assert "precio unitario final sin IVA" in respuesta.json()["detail"]
